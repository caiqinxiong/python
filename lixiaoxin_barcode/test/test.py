# -*- coding: utf-8 -*-
__author__ = 'caiqinxiong_cai'
#2019/12/17 10:44
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from pystrich.code128 import Code128Encoder


def excelCode128(file):
    '''
    file：在文件目录下执行，excel的文件名。
    A列是商品编码，B列是生成的条形码。
    '''

    wb = load_workbook(file)
    sheet = wb.active

    x = 0

    for r in range(1, sheet.max_row + 1):
        for i in range(1, 3):
            if i == 1:
                # A列数据生成条码
                var = sheet.cell(column=i, row=r).value
                filename = str(var) + '.PNG'
                Code128Encoder(str(var)).save(filename, bar_width=1)
            if i == 2:
                # B列数据插入条码
                img = Image(filename)
                B = 'B' + str(r)
                sheet.add_image(img, B)
        x += 1

    wb.save(file)

    print("执行成功！共计%s条！" % str(x))

excelCode128(r'E:\python_test\python\lixiaoxin_barcode\aa.xlsx')
